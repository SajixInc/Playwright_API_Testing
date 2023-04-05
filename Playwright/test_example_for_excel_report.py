import datetime
import json

from playwright.sync_api import APIRequestContext,Playwright
from typing import Generator
import pytest
from openpyxl import load_workbook


EnvironmentName = "https://jsonplaceholder.typicode.com"
Version = "v3"

@pytest.fixture(scope="session")
def api_request_context(
    playwright: Playwright,
)-> Generator[APIRequestContext,None,None]:
    request_context = playwright.request.new_context()
    yield request_context
    request_context.dispose()




def test_post_example(api_request_context: APIRequestContext) -> None:
    wb = load_workbook("Report.xlsx", data_only=True)
    sh = wb["API"]
    d = sh.max_row
    endpoint = 'https://jsonplaceholder.typicode.com/posts'
    headers = {"Content-type": "application/json"}
    data = {
        "title": "foo",
        "body": "bar",
        "userId": 1
    }
    post_todo = api_request_context.post(
        f"https://jsonplaceholder.typicode.com/posts", data=data, headers=headers
    )
    # assert post_todo.ok
    # assert post_todo.json()["title"] == "foo"
    test1 = post_todo.json()
    if test1['title'] == 'foo':
        sh[f'A{d + 1}'].value = '1'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        # sh[f'E{d + 1}'].value = test1['Status']
        sh[f'E{d + 1}'].value = 200
        sh[f'F{d + 1}'].value = "Success"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    else:
        sh[f'A{d + 1}'].value = '1'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 400
        sh[f'F{d + 1}'].value = "Fail"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    wb.save("Report.xlsx")
    wb.close()
    assert post_todo.json()["title"] == "foo"



def test_get_example(api_request_context: APIRequestContext):
    headers = {"Content-type": "application/json"}
    wb = load_workbook("Report.xlsx", data_only=True)
    sh = wb["API"]
    d = sh.max_row
    endpoint = 'https://jsonplaceholder.typicode.com/posts/getbyid'
    ### In place of getbyid you have to give which id details needs tobe get
    get_todo = api_request_context.get(
        f"https://jsonplaceholder.typicode.com/posts/1",  headers=headers
    )

    test1 = get_todo.json()
    if test1['userId'] == 1:
        sh[f'A{d + 1}'].value = '2'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 200
        sh[f'F{d + 1}'].value = "Success"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    else:
        sh[f'A{d + 1}'].value = '2'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 400
        sh[f'F{d + 1}'].value = "Fail"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    wb.save("Report.xlsx")
    wb.close()
    assert get_todo.ok


def test_put_example(api_request_context: APIRequestContext):
    headers = {"Content-type": "application/json"}
    wb = load_workbook("Report.xlsx", data_only=True)
    sh = wb["API"]
    d = sh.max_row
    endpoint = 'https://jsonplaceholder.typicode.com/posts/updatebyid'
    ### In place of updatebyid you have to give which id needs tobe update
    data = {
        "title": "foo2",
        "body": "bar",
        "userId": 1
    }
    put_todo = api_request_context.put(
        f"https://jsonplaceholder.typicode.com/posts/1", data=data, headers=headers
    )
    # assert put_todo.ok

    test1 = put_todo.json()
    if test1['title'] == 'foo2':
        sh[f'A{d + 1}'].value = '3'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 200
        sh[f'F{d + 1}'].value = "Success"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    else:
        sh[f'A{d + 1}'].value = '3'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 400
        sh[f'F{d + 1}'].value = "Fail"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    wb.save("Report.xlsx")
    wb.close()
    assert put_todo.ok


def test_delete_example(api_request_context: APIRequestContext):
    headers = {"Content-type": "application/json"}
    wb = load_workbook("Report.xlsx", data_only=True)
    sh = wb["API"]
    d = sh.max_row
    endpoint = 'https://jsonplaceholder.typicode.com/posts/deletebyid'

    ### In place of deletebyid you have to give which id needs tobe delete

    delete_todo = api_request_context.delete(
        f"https://jsonplaceholder.typicode.com/posts/1", headers=headers)

    test1 = delete_todo.json()
    print(test1)
    if test1 == {}:
        sh[f'A{d + 1}'].value = '4'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 200
        sh[f'F{d + 1}'].value = "Success"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
        sh[f'I{d + 1}'].value = str(EnvironmentName)
    else:
        sh[f'A{d + 1}'].value = '4'
        sh[f'B{d + 1}'].value = str(endpoint)
        sh[f'C{d + 1}'].value = 'User'
        sh[f'D{d + 1}'].value = "User"
        sh[f'E{d + 1}'].value = 400
        sh[f'F{d + 1}'].value = "Fail"
        sh[f'G{d + 1}'].value = str(datetime.datetime.now())
        sh[f'H{d + 1}'].value = str(Version)
    wb.save("Report.xlsx")
    wb.close()
    assert delete_todo.ok

